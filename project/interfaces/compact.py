import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def consolidar_e_salvar(folhaDirectory, arquivo_saida):
    print("Executando...")
    responseObjects = []

    for arquivo in os.listdir(folhaDirectory):
        caminho_completo = os.path.join(folhaDirectory, arquivo)
        if arquivo.endswith('.xls') or arquivo.endswith('.xlsx'):
            engine = 'openpyxl' if arquivo.endswith('.xlsx') else None
            try:
                df = pd.read_excel(caminho_completo, engine=engine)
                responseObjects.append(df)
            except Exception as e:
                print(f"Erro ao ler o arquivo {arquivo}: {e}")
        else:
            continue

    if responseObjects:
        df_consolidado = pd.concat(responseObjects, ignore_index=True)

        # Cria uma nova pasta de trabalho do Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        max_rows_per_sheet = 1048576
        sheet_count = 1

        for r in dataframe_to_rows(df_consolidado, index=False, header=True):
            if ws.max_row >= max_rows_per_sheet:
                # Cria uma nova aba se a atual atingiu o limite
                sheet_count += 1
                ws = wb.create_sheet(title=f"Sheet{sheet_count}")
            ws.append(r)

        # Salva a pasta de trabalho
        wb.save(f"{arquivo_saida}.xlsx")

        print("Encerrou o processo...")
        print(f"Arquivo consolidado salvo como: {arquivo_saida}.xlsx")
    else:
        print("Nenhum arquivo Excel válido foi encontrado para processamento.")

# Definir o diretório e o nome do arquivo de saída
folhaDirectory = "C:/Users/localuser/Documents/joao/Planilhas/TEMP PLAN BOT"
arquivo_saida = "VRteste"

consolidar_e_salvar(folhaDirectory, arquivo_saida)
